from tkinter import Tk, ttk, Button
from tkinter import filedialog
from click import progressbar
import openpyxl
import pymysql
import requests
from bs4 import BeautifulSoup
import tqdm
from main import*
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
import pymysql
import tqdm
from tkinter import *
from tkinter import ttk
import os
import openpyxl
from tkinter import filedialog
import pymysql


i = 0
registros = []


def consulta2(cedula):
    registro = {}
    url = f"http://www.cne.gob.ve/web/registro_electoral/ce.php?nacionalidad=V&cedula={cedula}"
    response = requests.get(url)
    html = BeautifulSoup(response.text, 'html.parser')
    b = html.find_all('b')
    td = html.find_all('td')
    existe = True
    encontrado = False

    for num, item in enumerate(td):
        if num == 11 and item.text[0] == 'V':
            existe = True
        elif b[0].text == "Esta cédula de identidad no se encuentra inscrito en el Registro Electoral.":
            existe = False
            registro["estatus"] = "No registrado"
            registro["cedula"] = cedula
            guardar_registros(registro)
            return

        if existe:
            if num == 11:
                registro["cedula"] = item.text
            if num == 13:
                if item.text == "ESTATUS":
                    encontrado = True
                    break
                else:
                    registro["nombre"] = item.text

            if num == 15:
                registro["estado"] = item.text
            if num == 17:
                registro["municipio"] = item.text
            if num == 19:
                registro["parroquia"] = item.text

            if num == 21:
                registro["centro"] = item.text
            if num == 23:
                registro["direccion"] = item.text

                guardar_registros(registro)
                return

    if encontrado:
        for num, item in enumerate(td):
            if num == 11:
                registro["cedula"] = item.text
            if num == 20:
                registro["estatus"] = item.text
                guardar_registros(registro)
                return
# Función para exportar las cédulas


def importar_data(ventana):
    # Abrir el diálogo para seleccionar el archivo Excel
    filepath = filedialog.askopenfilename(
        title="Seleccionar archivo Excel", filetypes=[("Archivos Excel", "*.xlsx")])

    if filepath:
       # Configuración de la conexión a la base de datos
        host = 'localhost'
        user = 'root'
        password = ''
        database_name = 'Data'
        table_name = 'cedulas'

    else:
        return
    # Establecer conexión con el servidor MySQL
    connection = pymysql.connect(
        host=host,
        user=user,
        password=password,
        database=database_name
    )

    # Crear una instancia del cursor
    cursor = connection.cursor()

    # Leer el archivo Excel y extraer los datos de la primera columna
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active

    # Recorrer las celdas de la primera columna y realizar la inserción en la tabla
    total_celdas = len(sheet['A'])

    for index, cell in enumerate(sheet['A']):
        cedula = cell.value

        # Verificar si la cédula ya existe en la base de datos
        select_query = f"SELECT cedula FROM {table_name} WHERE cedula = %s"
        cursor.execute(select_query, (cedula,))
        result = cursor.fetchone()

        # Si la cédula no existe, realizar la inserción
        if not result:
            insert_query = f"INSERT INTO {table_name} (cedula) VALUES (%s)"
            cursor.execute(insert_query, (cedula,))

          # Confirmar los cambios en la base de datos
            connection.commit()
        progreso_actual = (index + 1) / total_celdas * 100

        # Actualizar la barra de progreso de la ventana
        ventana.progress_bar['value'] = progreso_actual
        ventana.progress_bar.update()


def consulta(ventana):
    host = 'localhost'
    user = 'root'
    password = ''
    database_name = 'Data'
    table_name = 'cedulas'

    # Establecer conexión con el servidor MySQL
    connection = pymysql.connect(
        host=host,
        user=user,
        password=password,
        database=database_name
    )
    cursor = connection.cursor()
    count_query = f"SELECT COUNT(cedula) FROM {table_name}"
    cursor.execute(count_query)
    total_cedulas = cursor.fetchone()[0]

    # Definir el tamaño del lote de cédulas a consultar en cada iteración
    lote_size = 100

    # Calcular el número total de lotes
    total_lotes = total_cedulas // lote_size + (total_cedulas % lote_size > 0)

    cedulas_procesadas = 0

    # Consultar las cédulas por lotes
    for index in range(total_lotes):
        offset = index * lote_size
        select_query = f"SELECT cedula FROM {table_name} LIMIT {lote_size} OFFSET {offset}"
        cursor.execute(select_query)

        # Iterar sobre las cédulas del lote y llamar a la función consulta2(cedula)
        for row in cursor.fetchall():
            cedula = row[0]
            consulta2(cedula)

            # Incrementar el contador de la barra de progreso

            cedulas_procesadas += 1
            progreso_actual = int(cedulas_procesadas / total_cedulas * 100)

            # Actualizar el campo de texto con el progreso en números
            ventana.progreso_text.delete(0, END)
            ventana.progreso_text.insert(
                0, f"{cedulas_procesadas}/{total_cedulas}")
            ventana.progreso_text.update()

        # Actualizar la barra de progreso de la ventana
            ventana.progress_bar['value'] = progreso_actual
            ventana.progress_bar.update()


def guardar_registros(registro):
    global i
    # Configuración de la conexión a la base de datos
    host = 'localhost'
    user = 'root'
    password = ''
    database_name = 'Data'
    table_name = 'Registros'

    # Establecer conexión con el servidor MySQL
    connection = pymysql.connect(
        host=host,
        user=user,
        password=password,
        database=database_name
    )

    # Crear una instancia del cursor
    cursor = connection.cursor()

    # Insertar el registro en la nueva tabla
    # Agregar este print para verificar el contenido del diccionario
    if "estatus" in registro:
        cedula = registro["cedula"]
        estatus = registro["estatus"]

        # Consultar si la cédula ya existe en la nueva tabla
        select_query = f"SELECT cedula FROM {table_name} WHERE cedula = %s"
        cursor.execute(select_query, (cedula,))
        result = cursor.fetchone()

        # Si la cédula no existe, realizar la inserción
        if not result:
            insert_query = f"INSERT INTO {table_name} (cedula, estatus) VALUES (%s, %s)"
            cursor.execute(insert_query, (cedula, estatus))

    else:
        cedula = registro['cedula']
        estatus = "Registrada"
        nombre = registro.get('nombre', '')
        estado = registro.get('estado', '')
        municipio = registro.get('municipio', '')
        parroquia = registro.get('parroquia', '')
        centro = registro.get('centro', '')
        direccion = registro.get('direccion', '')

        # Consultar si la cédula ya existe en la nueva tabla
        select_query = f"SELECT cedula FROM {table_name} WHERE cedula = %s"
        cursor.execute(select_query, (cedula,))
        result = cursor.fetchone()

        # Si la cédula no existe, realizar la inserción
        if not result:
            insert_query = f"INSERT INTO {table_name} (cedula, estatus, nombre, estado, municipio, parroquia, centro, direccion) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
            cursor.execute(insert_query, (cedula, estatus, nombre,
                                          estado, municipio, parroquia, centro, direccion))

    connection.commit()

    # Recorrer los resultados y mostrar cada fila en la consola


def exportar_a_excel():
    # Configuración de la conexión a la base de datos
    host = 'localhost'
    user = 'root'
    password = ''
    database_name = 'Data'
    table_name = 'Registros'

    # Establecer conexión con el servidor MySQL
    connection = pymysql.connect(
        host=host,
        user=user,
        password=password,
        database=database_name
    )

    # Crear una instancia del cursor
    cursor = connection.cursor()

    # Obtener todos los registros de la tabla "Registros"
    select_query = f"SELECT * FROM {table_name}"
    cursor.execute(select_query)
    registros = cursor.fetchall()

    # Crear el diálogo para seleccionar o crear un archivo
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Archivos de Excel", "*.xlsx")],
        title="Guardar consulta"
    )

    if file_path:
        # Crear un nuevo libro de Excel y seleccionar la hoja activa
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Obtener los nombres de las columnas de la tabla
        column_names = [column[0] for column in cursor.description]

        # Escribir los nombres de las columnas en la primera fila del archivo Excel
        for col_idx, column_name in enumerate(column_names, 1):
            sheet.cell(row=1, column=col_idx).value = column_name

        # Escribir los datos de los registros en las filas siguientes
        for row_idx, registro in enumerate(registros, 2):
            for col_idx, valor in enumerate(registro, 1):
                sheet.cell(row=row_idx, column=col_idx).value = valor

        # Ajustar el ancho de las columnas según el contenido del texto
        for column in sheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Guardar el archivo Excel en la ubicación seleccionada por el usuario
        workbook.save(file_path)
        os.startfile(file_path)


def vaciar_tabla(nombre):
    # Configuración de la conexión a la base de datos
    host = 'localhost'
    user = 'root'
    password = ''
    database_name = 'Data'
    table_name = nombre

    # Establecer conexión con el servidor MySQL
    connection = pymysql.connect(
        host=host,
        user=user,
        password=password,
        database=database_name
    )

    # Crear una instancia del cursor
    cursor = connection.cursor()

    # Ejecutar la sentencia DELETE para vaciar la tabla
    delete_query = f"DELETE FROM {table_name}"
    cursor.execute(delete_query)

    # Confirmar los cambios en la base de datos
    connection.commit()
